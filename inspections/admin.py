from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django import forms
from .models import Inspection, DefectDetail, InspectionInput, MasterData, AssemblingInspection, AssemblingDefectDetail, AssemblingInspectionInput, AssemblingMasterData
import openpyxl
from datetime import datetime
from django.utils import timezone
from django_admin_listfilter_dropdown.filters import DropdownFilter, ChoiceDropdownFilter
from rangefilter.filters import DateRangeFilter
from django.db import transaction

def natural_sort_key(item):
    """Sort key that sorts numeric strings numerically, text alphabetically after numbers."""
    val = item[0]  # item is a (value, label) tuple
    try:
        return (0, float(val), val)  # numeric: sort numerically first
    except (ValueError, TypeError):
        return (1, 0, str(val).lower())  # non-numeric: sort alphabetically after numbers

class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(label="Upload Excel File")

class InspectionForm(forms.ModelForm):
    TIME_CHOICES = [
        ('', '---------'),
        ('06.30-07.30', '06.30-07.30'),
        ('07.30-08.30', '07.30-08.30'),
        ('08.30-09.30', '08.30-09.30'),
        ('09.30-10.30', '09.30-10.30'),
        ('10.30-11.30', '10.30-11.30'),
        ('11.30-11.45', '11.30-11.45'),
        ('12.45-13.45', '12.45-13.45'),
        ('13.45-14.45', '13.45-14.45'),
        ('14.45-15.45', '14.45-15.45'),
        ('15.45-16.45', '15.45-16.45'),
        ('16.45-17.45', '16.45-17.45'),
        ('17.45-18.00', '17.45-18.00'),
        ('18.45-19.45', '18.45-19.45'),
        ('19.45-20.45', '19.45-20.45'),
        ('20.45-21.45', '20.45-21.45'),
        ('21.45-22.45', '21.45-22.45'),
        ('22.45-23.45', '22.45-23.45'),
        ('23.45-00.45', '23.45-00.45'),
    ]

    time_range = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_time', 'autocomplete': 'off'}), required=False)
    line = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_line', 'autocomplete': 'off'}), required=False)
    qa_name = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_qa_name', 'autocomplete': 'off'}), required=False)
    qa_id = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_qa_id', 'autocomplete': 'off'}), required=False)
    brand = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_brand', 'autocomplete': 'off'}), required=False)
    model = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_model', 'autocomplete': 'off'}), required=False)
    size = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_size', 'autocomplete': 'off'}), required=False)
    colour = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_colour', 'autocomplete': 'off'}), required=False)
    item_name = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_item', 'autocomplete': 'off'}), required=False)

    class Meta:
        model = Inspection
        fields = '__all__'

@admin.register(MasterData)
class MasterDataAdmin(admin.ModelAdmin):
    list_display = ('category', 'value')  # type: ignore
    list_filter = ('category',)
    search_fields = ('value',)

@admin.register(AssemblingMasterData)
class AssemblingMasterDataAdmin(admin.ModelAdmin):
    list_display = ('category', 'value')  # type: ignore
    list_filter = ('category',)
    search_fields = ('value',)

class DefectDetailInline(admin.TabularInline):
    model = DefectDetail
    extra = 1

@admin.register(Inspection)
class InspectionAdmin(admin.ModelAdmin):
    form = InspectionForm
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
    list_display = (  # type: ignore
        'date', 'time_range', 'qa_name', 'qa_id', 'shift', 
        'line', 'brand', 'model', 'size', 'colour', 
        'po_no', 'item_name', 'production_output', 'qty_check', 'get_defects'
    )
    
    fieldsets = (
        (None, {
            'classes': ('horizontal-fields',),
            'fields': (
                ('date', 'time_range', 'shift', 'line', 'qa_name', 'qa_id'),
                ('brand', 'model', 'size', 'colour', 'po_no', 'item_name'),
                ('production_output', 'qty_check'),
            )
        }),
    )
    list_filter = (
        ('date', DateRangeFilter),
        ('shift', ChoiceDropdownFilter),
        ('brand', DropdownFilter),
        ('qa_name', DropdownFilter),
    )
    search_fields = ('item_name', 'po_no', 'qa_name')
    inlines = [DefectDetailInline]  # type: ignore
    
    change_list_template = "admin/inspections_changelist.html"

    def has_add_permission(self, request):
        """Remove Add button from the main table so users use the specific Input menu."""
        return False

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        master_qs = MasterData.objects.all()
        categories = {
            'TIME': [choice[0] for choice in InspectionForm.TIME_CHOICES if choice[0]],
            'LINE': [], 'QA_NAME': [], 'QA_ID': [], 'BRAND': [],
            'MODEL': [], 'SIZE': [], 'COLOUR': [], 'ITEM': []
        }
        for md in master_qs:
            if md.category in categories and md.value not in categories[md.category]:
                categories[md.category].append(md.value)
        for cat in categories:
            if cat == 'LINE':
                categories[cat] = sorted(categories[cat], key=lambda x: (0, float(x), x) if x.replace('.','',1).isdigit() else (1, 0, str(x).lower()))
            elif cat != 'TIME':
                categories[cat] = sorted(categories[cat], key=lambda x: str(x).lower())
        context['master_data_lists'] = categories
        return super().render_change_form(request, context, add, change, form_url, obj)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name="inspections_inspection_import_excel")
        ]
        return custom_urls + urls

    def get_defects(self, obj):
        defects = obj.defects.all()
        if not defects:
            return "-"
        return ", ".join([f"{d.defect_type} ({d.qty})" for d in defects])
    get_defects.short_description = "Defects"  # type: ignore

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    ws = wb.active
                    
                    headers = [cell.value for cell in ws[1]]  # type: ignore
                    success_count = 0
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore
                        if not any(row): continue
                        
                        row_dict = dict(zip(headers, row))
                        
                        date_val = row_dict.get("DATE")
                        if isinstance(date_val, datetime):
                            date_str = date_val.date()
                        else:
                            date_str = str(date_val).split(" ")[0] if date_val else datetime.now().date()
                        
                        inspection, created = Inspection.objects.get_or_create(
                            date=date_str,
                            time_range=str(row_dict.get("Time") or "08.00-09.00"),
                            qa_name=str(row_dict.get("QA name") or "Unknown"),
                            qa_id=str(row_dict.get("QA ID") or "Unknown"),
                            shift=str(row_dict.get("Shift") or "1"),
                            line=str(row_dict.get("Line") or "Line 1"),
                            brand=str(row_dict.get("Brand") or "Unknown"),
                            model=str(row_dict.get("Model") or "Unknown"),
                            size=str(row_dict.get("Size") or "Unknown"),
                            colour=str(row_dict.get("Colour") or "Unknown"),
                            po_no=str(row_dict.get("Po.No.") or "Unknown"),
                            item_name=str(row_dict.get("Item Name") or "Unknown"),
                            defaults={
                                'production_output': int(row_dict.get("Production Output") or 0),
                                'qty_check': int(row_dict.get("Qty Check") or 0)
                            }
                        )
                        
                        reject_type = row_dict.get("Reject Types")
                        if reject_type:
                            qty_reject = int(row_dict.get("Qty Reject") or 1)
                            repair_method = str(row_dict.get("How to Repair") or "None")
                            
                            DefectDetail.objects.create(
                                inspection=inspection,
                                defect_type=str(reject_type),
                                repair_method=repair_method,
                                qty=qty_reject
                            )
                        
                        success_count += 1
                        
                    self.message_user(request, f"Successfully imported {success_count} rows from Excel.")
                    return redirect("..")
                except Exception as e:
                    self.message_user(request, f"Error processing file: {str(e)}", level=messages.ERROR)
                    return redirect("..")
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'opts': self.model._meta,
            'title': 'Import Inspections from Excel',
        }
        return render(request, "admin/excel_import.html", context)

@admin.register(DefectDetail)
class DefectDetailAdmin(admin.ModelAdmin):
    list_display = ('inspection', 'defect_type', 'qty', 'repair_method')  # type: ignore
    list_filter = ('defect_type',)
    search_fields = ('defect_type', 'repair_method')


@admin.register(InspectionInput)
class InspectionInputAdmin(admin.ModelAdmin):
    form = InspectionForm
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
    """
    Specific menu for INPUT data.
    We keep ADD permission active here.
    """
    list_display = ('item_name', 'date', 'qa_name', 'shift')  # type: ignore
    inlines = [DefectDetailInline]  # type: ignore
    fieldsets = InspectionAdmin.fieldsets
    change_form_template = "admin/inspections_input_change_form.html"
    
    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        today = timezone.now().date()
        recent_inputs = Inspection.objects.filter(created_at__date=today).order_by('-created_at')[:15]
        recent_inputs = recent_inputs.prefetch_related('defects')
        context['recent_inputs'] = recent_inputs
        
        master_qs = MasterData.objects.all()
        categories = {
            'TIME': [choice[0] for choice in InspectionForm.TIME_CHOICES if choice[0]],
            'LINE': [], 'QA_NAME': [], 'QA_ID': [], 'BRAND': [],
            'MODEL': [], 'SIZE': [], 'COLOUR': [], 'ITEM': []
        }
        for md in master_qs:
            if md.category in categories and md.value not in categories[md.category]:
                categories[md.category].append(md.value)
        for cat in categories:
            if cat == 'LINE':
                categories[cat] = sorted(categories[cat], key=lambda x: (0, float(x), x) if x.replace('.','',1).isdigit() else (1, 0, str(x).lower()))
            elif cat != 'TIME':
                categories[cat] = sorted(categories[cat], key=lambda x: str(x).lower())
        context['master_data_lists'] = categories
        return super().render_change_form(request, context, add, change, form_url, obj)
    
    def has_module_permission(self, request):
        return True

    def changelist_view(self, request, extra_context=None):
        return redirect('admin:inspections_inspectioninput_add')

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name="inspections_inspectioninput_import_excel")
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    ws = wb.active
                    
                    headers = [cell.value for cell in ws[1]]  # type: ignore
                    
                    # POKA YOKE Validation
                    if "Inspection Quantity" in headers or "Qty Check Hours" in headers:
                        self.message_user(request, "ERROR: File ditolak! Anda mencoba meng-upload file Assembling ke dalam form Sewing.", level=messages.ERROR)
                        return redirect("..")
                    
                    success_count = 0
                    
                    for row in ws.iter_rows(min_row=2, values_only=True):  # type: ignore
                        if not any(row): continue
                        
                        row_dict = dict(zip(headers, row))
                        
                        date_val = row_dict.get("DATE")
                        if isinstance(date_val, datetime):
                            date_str = date_val.date()
                        else:
                            date_str = str(date_val).split(" ")[0] if date_val else datetime.now().date()
                        
                        inspection, created = Inspection.objects.get_or_create(
                            date=date_str,
                            time_range=str(row_dict.get("Time") or "08.00-09.00"),
                            qa_name=str(row_dict.get("QA name") or "Unknown"),
                            qa_id=str(row_dict.get("QA ID") or "Unknown"),
                            shift=str(row_dict.get("Shift") or "1"),
                            line=str(row_dict.get("Line") or "Line 1"),
                            brand=str(row_dict.get("Brand") or "Unknown"),
                            model=str(row_dict.get("Model") or "Unknown"),
                            size=str(row_dict.get("Size") or "Unknown"),
                            colour=str(row_dict.get("Colour") or "Unknown"),
                            po_no=str(row_dict.get("Po.No.") or "Unknown"),
                            item_name=str(row_dict.get("Item Name") or "Unknown"),
                            defaults={
                                'production_output': int(row_dict.get("Production Output") or 0),
                                'qty_check': int(row_dict.get("Qty Check") or 0)
                            }
                        )
                        
                        reject_type = row_dict.get("Reject Types")
                        if reject_type:
                            qty_reject = int(row_dict.get("Qty Reject") or 1)
                            repair_method = str(row_dict.get("How to Repair") or "None")
                            
                            DefectDetail.objects.create(
                                inspection=inspection,
                                defect_type=str(reject_type),
                                repair_method=repair_method,
                                qty=qty_reject
                            )
                        
                        # Auto-populate MasterData from imported values
                        for cat, key_name in [
                            ('LINE', 'Line'),
                            ('QA_NAME', 'QA name'),
                            ('QA_ID', 'QA ID'),
                            ('BRAND', 'Brand'),
                            ('MODEL', 'Model'),
                            ('SIZE', 'Size'),
                            ('COLOUR', 'Colour'),
                            ('ITEM', 'Item Name'),
                        ]:
                            val = row_dict.get(key_name)
                            if val and str(val).strip() != "" and str(val).strip().lower() not in ["unknown", "none", "nan"]:
                                MasterData.objects.get_or_create(
                                    category=cat,
                                    value=str(val).strip()
                                )
                        
                        success_count += 1
                        
                    self.message_user(request, f"Successfully imported {success_count} rows from Excel.")
                    return redirect("..")
                except Exception as e:
                    self.message_user(request, f"Error processing file: {str(e)}", level=messages.ERROR)
                    return redirect("..")
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'opts': self.model._meta,
            'title': 'Import Sewing Inspections from Excel',
        }
        return render(request, "admin/excel_import.html", context)


class AssemblingDefectDetailInline(admin.TabularInline):
    model = AssemblingDefectDetail
    extra = 1

class AssemblingInspectionForm(forms.ModelForm):
    line = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_line', 'autocomplete': 'off'}), required=False)
    qa_name = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_qa_name', 'autocomplete': 'off'}), required=False)
    qa_id = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_qa_id', 'autocomplete': 'off'}), required=False)
    brand = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_brand', 'autocomplete': 'off'}), required=False)
    model = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_model', 'autocomplete': 'off'}), required=False)
    size = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_size', 'autocomplete': 'off'}), required=False)
    colour = forms.CharField(widget=forms.TextInput(attrs={'list': 'datalist_colour', 'autocomplete': 'off'}), required=False)

    class Meta:
        model = AssemblingInspection
        fields = '__all__'

@admin.register(AssemblingInspection)
class AssemblingInspectionAdmin(admin.ModelAdmin):
    form = AssemblingInspectionForm
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
    list_display = (  # type: ignore
        'date', 'qa_name', 'qa_id', 'shift', 
        'line', 'brand', 'model', 'size', 'colour', 
        'po_no', 'inspection_quantity', 'qty_check_hours', 'get_defects'
    )
    
    fieldsets = (
        (None, {
            'classes': ('horizontal-fields',),
            'fields': (
                ('date', 'shift', 'line', 'qa_name', 'qa_id'),
                ('brand', 'model', 'size', 'colour', 'po_no'),
                ('inspection_quantity', 'qty_check_hours'),
            )
        }),
    )
    list_filter = (
        ('date', DateRangeFilter),
        ('shift', ChoiceDropdownFilter),
        ('brand', DropdownFilter),
        ('qa_name', DropdownFilter),
    )
    search_fields = ('po_no', 'qa_name')
    inlines = [AssemblingDefectDetailInline]  # type: ignore
    
    change_list_template = "admin/inspections_changelist.html"

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        total_defects = sum([defect.qty for defect in obj.defects.all()])
        obj.defect_quantity = total_defects
        if obj.inspection_quantity and obj.inspection_quantity > 0:
            obj.defect_rate = total_defects / obj.inspection_quantity
        else:
            obj.defect_rate = 0.0
        obj.save()

    def has_add_permission(self, request):
        return False

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        master_qs = AssemblingMasterData.objects.all()
        categories = {
            'LINE': [], 'QA_NAME': [], 'QA_ID': [], 'BRAND': [],
            'MODEL': [], 'SIZE': [], 'COLOUR': []
        }
        for md in master_qs:
            if md.category in categories and md.value not in categories[md.category]:
                categories[md.category].append(md.value)
        for cat in categories:
            if cat == 'LINE':
                categories[cat] = sorted(categories[cat], key=lambda x: (0, float(x), x) if x.replace('.','',1).isdigit() else (1, 0, str(x).lower()))
            else:
                categories[cat] = sorted(categories[cat], key=lambda x: str(x).lower())
        context['master_data_lists'] = categories
        return super().render_change_form(request, context, add, change, form_url, obj)

    def get_defects(self, obj):
        defects = obj.defects.all()
        if not defects:
            return "-"
        return ", ".join([f"{d.defect_type} ({d.qty})" for d in defects])
    get_defects.short_description = "Defects"  # type: ignore

@admin.register(AssemblingInspectionInput)
class AssemblingInspectionInputAdmin(admin.ModelAdmin):
    form = AssemblingInspectionForm
    class Media:
        css = {
            'all': ('admin/css/custom_admin.css',)
        }
    list_display = ('po_no', 'date', 'qa_name', 'shift')  # type: ignore
    inlines = [AssemblingDefectDetailInline]  # type: ignore
    fieldsets = AssemblingInspectionAdmin.fieldsets
    change_form_template = "admin/assembling_input_change_form.html"

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        obj = form.instance
        total_defects = sum([defect.qty for defect in obj.defects.all()])
        obj.defect_quantity = total_defects
        if obj.inspection_quantity and obj.inspection_quantity > 0:
            obj.defect_rate = total_defects / obj.inspection_quantity
        else:
            obj.defect_rate = 0.0
        obj.save()
    
    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        today = timezone.now().date()
        recent_inputs = AssemblingInspection.objects.filter(created_at__date=today).order_by('-created_at')[:15]
        recent_inputs = recent_inputs.prefetch_related('defects')
        context['recent_inputs'] = recent_inputs
        
        master_qs = AssemblingMasterData.objects.all()
        categories = {
            'LINE': [], 'QA_NAME': [], 'QA_ID': [], 'BRAND': [],
            'MODEL': [], 'SIZE': [], 'COLOUR': []
        }
        for md in master_qs:
            if md.category in categories and md.value not in categories[md.category]:
                categories[md.category].append(md.value)
        for cat in categories:
            if cat == 'LINE':
                categories[cat] = sorted(categories[cat], key=lambda x: (0, float(x), x) if x.replace('.','',1).isdigit() else (1, 0, str(x).lower()))
            else:
                categories[cat] = sorted(categories[cat], key=lambda x: str(x).lower())
        context['master_data_lists'] = categories
        return super().render_change_form(request, context, add, change, form_url, obj)
    
    def has_module_permission(self, request):
        return True

    def changelist_view(self, request, extra_context=None):
        return redirect('admin:inspections_assemblinginspectioninput_add')

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name="inspections_assemblinginspectioninput_import_excel")
        ]
        return custom_urls + urls

    def import_excel(self, request):
        if request.method == "POST":
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                try:
                    wb = openpyxl.load_workbook(excel_file, data_only=True)
                    ws = wb.active
                    
                    # Find header row dynamically
                    header_row_idx = 1
                    headers_original = []
                    headers_lower = []
                    
                    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):  # type: ignore
                        row_strs = [str(v).strip().lower() for v in row if v is not None]
                        # Look for common header keywords
                        if any(k in row_strs for k in ['date', 'tanggal']) and any(k in row_strs for k in ['shift', 'line', 'qa name', 'brand']):
                            header_row_idx = row_idx
                            headers_original = [str(v).strip() if v is not None else "" for v in row]
                            headers_lower = [h.lower() for h in headers_original]
                            break
                            
                    if not headers_lower:
                        self.message_user(request, "ERROR: Tidak dapat menemukan baris Header (kolom DATE, SHIFT, dll). Pastikan format tabel benar.", level=messages.ERROR)
                        return redirect("..")
                    
                    # POKA YOKE Validation
                    if "production output" in headers_lower or "item name" in headers_lower or "how to repair" in headers_lower:
                        self.message_user(request, "ERROR: File ditolak! Anda mencoba meng-upload file Sewing ke dalam form Assembling.", level=messages.ERROR)
                        return redirect("..")

                    success_count = 0
                    empty_count = 0
                    
                    with transaction.atomic():  # type: ignore
                        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):  # type: ignore
                            if not any(row):
                                empty_count += 1
                                if empty_count >= 10:
                                    break
                                continue
                            empty_count = 0
                            
                            row_dict = dict(zip(headers_lower, row))
                            
                            def get_val(keys, default=None):
                                for k in keys:
                                    if k in row_dict and row_dict[k] is not None:
                                        return row_dict[k]
                                return default
                            
                            def safe_int(val, default=0):
                                try:
                                    if val is None or str(val).strip() == '': return default
                                    return int(float(val))
                                except (ValueError, TypeError):
                                    return default
                            
                            date_val = get_val(['date', 'tanggal'])
                            if isinstance(date_val, datetime):
                                date_str = date_val.date()
                            else:
                                date_str = str(date_val).split(" ")[0] if date_val else datetime.now().date()
                            
                            inspection, created = AssemblingInspection.objects.get_or_create(
                                date=date_str,
                                line=str(get_val(['line', 'jalur'], "Line 1")),
                                shift=str(get_val(['shift'], "1")),
                                qa_name=str(get_val(['qa name', 'qa_name', 'nama qa'], "Unknown")),
                                qa_id=str(get_val(['qa id', 'qa_id', 'nik', 'id qa'], "Unknown")),
                                po_no=str(get_val(['po no', 'po.no.', 'po_no', 'po', 'po number'], "Unknown")),
                                brand=str(get_val(['brand', 'merek', 'merk'], "Unknown")),
                                model=str(get_val(['model', 'style'], "Unknown")),
                                size=str(get_val(['size', 'ukuran'], "Unknown")),
                                colour=str(get_val(['colour', 'color', 'warna'], "Unknown")),
                                defaults={
                                    'inspection_quantity': safe_int(get_val(['inspection quantity', 'qty inspection', 'total check'])),
                                    'qty_check_hours': safe_int(get_val(['qty check hours', 'qty check/hours', 'qty check per hours']))
                                }
                            )
                            
                            # Auto-populate AssemblingMasterData from imported values
                            for cat, keys in [
                                ('LINE', ['line', 'jalur']),
                                ('QA_NAME', ['qa name', 'qa_name', 'nama qa']),
                                ('QA_ID', ['qa id', 'qa_id', 'nik', 'id qa']),
                                ('BRAND', ['brand', 'merek', 'merk']),
                                ('MODEL', ['model', 'style']),
                                ('SIZE', ['size', 'ukuran']),
                                ('COLOUR', ['colour', 'color', 'warna']),
                            ]:
                                val = get_val(keys)
                                if val and str(val).strip() != "" and str(val).strip().lower() not in ["unknown", "none", "nan"]:
                                    AssemblingMasterData.objects.get_or_create(
                                        category=cat,
                                        value=str(val).strip()
                                    )
                            
                            if not created:
                                inspection.defects.all().delete()  # type: ignore[attr-defined]
                                
                            base_columns_lower = [
                                'date', 'tanggal', 'line', 'jalur', 'shift', 'qa name', 'qa_name', 'nama qa', 
                                'qa id', 'qa_id', 'nik', 'id qa', 'po.no.', 'po no', 'po_no', 'po', 'po number', 
                                'brand', 'merek', 'merk', 'model', 'style', 'size', 'ukuran', 'colour', 'color', 'warna', 
                                'inspection quantity', 'qty inspection', 'total check',
                                'qty check/hours', 'qty check hours', 'qty check per hours', 'defect quantity', 'defect rate', 'defect rate (%)'
                            ]
                            
                            for i, col_val in enumerate(row):
                                if i >= len(headers_original): break
                                orig_col_name = headers_original[i]
                                lower_col_name = headers_lower[i]
                                
                                if lower_col_name and lower_col_name not in base_columns_lower and col_val is not None:
                                    try:
                                        qty_defect = int(col_val)
                                        if qty_defect > 0:
                                            AssemblingDefectDetail.objects.create(
                                                inspection=inspection,
                                                defect_type=str(orig_col_name).strip(),
                                                qty=qty_defect
                                            )
                                    except (ValueError, TypeError):
                                        pass
                            
                            # Update defect summary
                            total_defects = sum([defect.qty for defect in inspection.defects.all()])  # type: ignore[attr-defined]
                            inspection.defect_quantity = total_defects
                            if inspection.inspection_quantity and inspection.inspection_quantity > 0:
                                inspection.defect_rate = total_defects / inspection.inspection_quantity
                            else:
                                inspection.defect_rate = 0.0
                            inspection.save()
                            
                            success_count += 1
                        
                    self.message_user(request, f"Successfully imported {success_count} rows dari baris ke-{header_row_idx+1}.")
                    return redirect("..")
                except Exception as e:
                    self.message_user(request, f"Error processing file: {str(e)}", level=messages.ERROR)
                    return redirect("..")
        else:
            form = ExcelImportForm()

        context = {
            'form': form,
            'opts': self.model._meta,
            'title': 'Import Assembling Inspections from Excel',
        }
        return render(request, "admin/excel_import.html", context)

